from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.style import Color, Fill
from openpyxl.cell import Cell

FILE_NAME = 'template.xlsx'
workbook = load_workbook(FILE_NAME)
worksheet =  workbook.get_sheet_by_name("Sheet1")

result = (
    ['Ram', 96],
    ['Ravi',   2],
    ['Kush',  95],
    ['Shyam',    95],
)

row_index = 0
col_index = 0
sum = 0

for name, marks in (result):
    
    worksheet.cell(row = row_index, column = col_index).value = name
    worksheet.cell(row = row_index, column = col_index + 1).value = marks
    sum += marks
    row_index += 1

no_of_students = len(result)
average = float(sum)/no_of_students

background =  Color.DARKRED
if average > 50:
	background = Color.YELLOW

worksheet.cell(row = row_index, column = 0).value = 'Average'
worksheet.cell(row = row_index, column = 1).value = average
worksheet.cell(row = row_index, column = 1).style.fill.fill_type = Fill.FILL_SOLID
worksheet.cell(row = row_index, column = 1).style.fill.start_color.index =  background


workbook.save(FILE_NAME)